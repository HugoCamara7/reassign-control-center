"""Lectura y escritura de los Excel de pedidos.

Objetivo central: **no perder ni deformar nada** del archivo original.
El archivo real de Forus (`Formato de Carga Reasignacion.xls`) trae varias
trampas que este modulo resuelve de forma explicita:

* Es un `.xls` BIFF real (OLE2), no un `.xlsx` renombrado.
* El nombre de la hoja lleva un timestamp variable
  (`shipping_groups_1783611991147`), asi que nunca se busca por nombre.
* Encabezados con mojibake real (`N_Tracking_EnvÃ­o`, `Â¿Quien_Recibe?`).
  Se conservan byte a byte porque la plataforma destino los espera asi.
* IDs largos (`ShGroup`) guardados como `float64`: al imprimirlos Python
  usa notacion cientifica (`8.41824206260704e+17`). Se escriben como texto.
* `SKU`, `Unidades`, `Cod_Tienda_Asig` llegan como `float` (`5438957.0`).
* `Fecha_Compra` viene con comillas dobles literales dentro del valor.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any

import pandas as pd

# A partir de este numero de digitos, un entero deja de ser representable de
# forma exacta en un float de 64 bits: se escribe como texto para que ni
# Python ni Excel lo reescriban en notacion cientifica.
SAFE_INT_LIMIT = 10**15

_SCIENTIFIC = re.compile(r"^[+-]?\d+(\.\d+)?[eE][+-]?\d+$")

# `5438957.0` / `5438957.00` -> `5438957`. Un decimal real (`1.5`) no se toca.
_TRAILING_ZERO_DECIMALS = re.compile(r"\.0+$")


@dataclass
class WorkbookPayload:
    """Contenido de un archivo de pedidos ya normalizado."""

    df: pd.DataFrame
    headers: list[str]
    sheet_name: str
    source_name: str
    source_format: str  # "xls" | "xlsx"
    n_rows: int = 0
    notes: list[str] = field(default_factory=list)

    def __post_init__(self) -> None:
        self.n_rows = len(self.df)


# ---------------------------------------------------------------------------
# Normalizacion de valores
# ---------------------------------------------------------------------------
def normalize_cell(value: Any) -> Any:
    """Devuelve el valor con el tipo mas fiel posible al original.

    Los floats que en realidad son enteros vuelven a `int`; los enteros
    demasiado grandes para un float se convierten a `str` para que no se
    impriman jamas en notacion cientifica.
    """
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, bool):
        return value
    if isinstance(value, float):
        if pd.isna(value):
            return ""
        if value.is_integer():
            as_int = int(value)
            return str(as_int) if abs(as_int) >= SAFE_INT_LIMIT else as_int
        return value
    if isinstance(value, int):
        return str(value) if abs(value) >= SAFE_INT_LIMIT else value
    return value


def as_text(value: Any) -> str:
    """Representacion en texto limpia, sin `.0` ni notacion cientifica."""
    if value is None:
        return ""
    if isinstance(value, float):
        if pd.isna(value):
            return ""
        if value.is_integer():
            return f"{int(value)}"
        return f"{value}".strip()
    text = str(value).strip()
    if text.lower() in ("nan", "nat", "none"):
        return ""
    if _SCIENTIFIC.match(text):
        try:
            number = float(text)
            if number.is_integer():
                return f"{int(number)}"
        except (TypeError, ValueError):
            pass
    return text


def normalize_sku(value: Any) -> str:
    """SKU canonico: mayusculas, sin `.0` de sobra y sin ceros a la izquierda.

    Los ceros importan. El archivo de pedidos suele traer el SKU como texto
    (`0005438957`) mientras BigQuery lo guarda como numero (`5438957`): si cada
    lado conserva su forma el cruce falla y la app informa "sin stock" para un
    producto que si lo tiene. Solo se recortan los ceros de un codigo
    completamente numerico; `0A12` se deja intacto porque ahi el cero puede ser
    parte del codigo.

    La misma transformacion se aplica dentro de la consulta a BigQuery
    (`SKU_SQL_EXPR`), para que ambos lados comparen exactamente lo mismo.
    """
    text = _TRAILING_ZERO_DECIMALS.sub("", as_text(value).upper()).strip()
    if text.isdigit():
        text = text.lstrip("0") or "0"
    return text


def normalize_store_code(value: Any) -> str:
    """Codigo de tienda canonico (`151.0` -> `151`)."""
    text = as_text(value)
    if not text:
        return ""
    match = re.match(r"^0*(\d+)(?:\.0+)?$", text)
    return match.group(1) if match else text.upper()


def normalize_store_name(value: Any) -> str:
    """Nombre de tienda comparable: mayusculas y espacios colapsados."""
    text = as_text(value).upper()
    text = text.replace(".", " ").replace("-", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize_status(value: Any) -> str:
    """`sin stock`, `Sin-Stock`, `SIN_STOCK` -> `SIN_STOCK`."""
    text = as_text(value).upper()
    text = re.sub(r"[\s\-]+", "_", text)
    return re.sub(r"_+", "_", text).strip("_")


def to_units(value: Any) -> int:
    """Unidades solicitadas como entero >= 0. Vacio/no numerico -> 0."""
    text = as_text(value).replace(",", ".")
    if not text:
        return 0
    try:
        return max(0, int(round(float(text))))
    except (TypeError, ValueError):
        return 0


def to_int(value: Any, default: int = 0) -> int:
    text = as_text(value).replace(",", ".")
    if not text:
        return default
    try:
        return int(round(float(text)))
    except (TypeError, ValueError):
        return default


# ---------------------------------------------------------------------------
# Lectura
# ---------------------------------------------------------------------------
def _dedupe_headers(headers: list[str]) -> tuple[list[str], list[str]]:
    """Evita columnas con el mismo nombre sin renombrar la primera aparicion."""
    seen: dict[str, int] = {}
    result: list[str] = []
    notes: list[str] = []
    for index, raw in enumerate(headers):
        name = str(raw).strip() if raw is not None else ""
        if not name:
            name = f"Columna_{index + 1}"
            notes.append(f"Columna {index + 1} sin encabezado: se nombro '{name}'.")
        if name in seen:
            seen[name] += 1
            duplicated = f"{name}__{seen[name]}"
            notes.append(f"Encabezado duplicado '{name}': se renombro a '{duplicated}'.")
            name = duplicated
        else:
            seen[name] = 0
        result.append(name)
    return result, notes


def _read_xls(content: bytes) -> tuple[list[str], list[list[Any]], str]:
    import xlrd

    book = xlrd.open_workbook(file_contents=content)
    sheet = book.sheet_by_index(0)
    if sheet.nrows == 0:
        return [], [], sheet.name

    headers = [sheet.cell_value(0, col) for col in range(sheet.ncols)]
    rows: list[list[Any]] = []
    for row_index in range(1, sheet.nrows):
        rows.append([sheet.cell_value(row_index, col) for col in range(sheet.ncols)])
    return headers, rows, sheet.name


def _read_xlsx(content: bytes) -> tuple[list[str], list[list[Any]], str]:
    from openpyxl import load_workbook

    book = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    sheet = book.worksheets[0]
    iterator = sheet.iter_rows(values_only=True)
    try:
        headers = list(next(iterator))
    except StopIteration:
        return [], [], sheet.title
    width = len(headers)
    rows = []
    for row in iterator:
        values = list(row)[:width]
        values += [None] * (width - len(values))
        rows.append(values)
    return [h if h is not None else "" for h in headers], rows, sheet.title


def read_orders(content: bytes, source_name: str) -> WorkbookPayload:
    """Lee un `.xls` o `.xlsx` de pedidos preservando encabezados y valores."""
    lowered = source_name.lower()
    if lowered.endswith(".xls"):
        source_format = "xls"
        headers, rows, sheet_name = _read_xls(content)
    elif lowered.endswith((".xlsx", ".xlsm")):
        source_format = "xlsx"
        headers, rows, sheet_name = _read_xlsx(content)
    else:
        # Sin extension reconocible: se decide por la firma del archivo.
        if content[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
            source_format = "xls"
            headers, rows, sheet_name = _read_xls(content)
        else:
            source_format = "xlsx"
            headers, rows, sheet_name = _read_xlsx(content)

    if not headers:
        raise ValueError("El archivo no tiene encabezados legibles en su primera hoja.")

    clean_headers, notes = _dedupe_headers([str(h) for h in headers])

    normalized_rows = [[normalize_cell(value) for value in row] for row in rows]
    df = pd.DataFrame(normalized_rows, columns=clean_headers, dtype=object)
    # Se descartan filas totalmente vacias (Excel suele arrastrar cientos).
    if not df.empty:
        mask = df.apply(lambda row: any(as_text(value) for value in row), axis=1)
        dropped = int((~mask).sum())
        if dropped:
            notes.append(f"Se ignoraron {dropped} filas completamente vacias.")
        df = df[mask].reset_index(drop=True)

    return WorkbookPayload(
        df=df,
        headers=clean_headers,
        sheet_name=sheet_name,
        source_name=source_name,
        source_format=source_format,
        notes=notes,
    )


def read_stock_file(content: bytes, source_name: str) -> pd.DataFrame:
    """Lee un Excel/CSV de stock manual (modo sin BigQuery).

    Se esperan al menos las columnas `sku`, `cod_tienda` y `stock`.
    """
    lowered = source_name.lower()
    if lowered.endswith(".csv"):
        df = pd.read_csv(io.BytesIO(content), dtype=object, sep=None, engine="python")
    else:
        payload = read_orders(content, source_name)
        df = payload.df
    df.columns = [str(column).strip() for column in df.columns]
    return df


# ---------------------------------------------------------------------------
# Escritura
# ---------------------------------------------------------------------------
def _write_value(cell, value: Any) -> None:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        cell.value = None
        return
    if isinstance(value, str):
        cell.value = value
        # Un texto que parece numero largo se fuerza a texto para que Excel
        # no lo convierta de vuelta a notacion cientifica.
        if value.isdigit() and len(value) > 11:
            cell.number_format = "@"
        return
    cell.value = value


def write_orders(
    df: pd.DataFrame,
    headers: list[str],
    sheet_name: str = "Reasignacion",
) -> bytes:
    """Genera el `.xlsx` final listo para cargar en la otra plataforma.

    `headers` fija el orden exacto de columnas de salida; cualquier columna
    del DataFrame que no este en la lista se omite (asi se pueden excluir
    las columnas de trazabilidad con un solo parametro).
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    book = Workbook(write_only=False)
    sheet = book.active
    # Excel no admite mas de 31 caracteres ni estos simbolos en el titulo.
    sheet.title = re.sub(r"[\[\]\*/\\\?:]", "-", sheet_name)[:31] or "Reasignacion"

    for index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.value = header

    for row_offset, (_, row) in enumerate(df.iterrows(), start=2):
        for col_index, header in enumerate(headers, start=1):
            _write_value(sheet.cell(row=row_offset, column=col_index), row.get(header, ""))

    # Anchos legibles sin recorrer todo el dataset.
    sample = df.head(200)
    for col_index, header in enumerate(headers, start=1):
        width = len(str(header))
        if header in sample.columns:
            for value in sample[header]:
                width = max(width, len(as_text(value)))
        sheet.column_dimensions[get_column_letter(col_index)].width = min(max(width + 2, 10), 46)

    sheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    book.save(buffer)
    return buffer.getvalue()


def write_report(sheets: dict[str, pd.DataFrame]) -> bytes:
    """Excel multi-hoja para el reporte operativo (KPIs, detalle, errores)."""
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            safe_name = re.sub(r"[\[\]\*/\\\?:]", "-", name)[:31] or "Hoja"
            body = frame if not frame.empty else pd.DataFrame({"Sin datos": []})
            body.to_excel(writer, sheet_name=safe_name, index=False)
    return buffer.getvalue()
